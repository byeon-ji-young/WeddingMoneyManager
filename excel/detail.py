# 엑셀 폴더 내 styles.py 불러오기
import excel.styles as styles

from openpyxl.utils import get_column_letter

def create_detail_sheet(worksheet, money_data, budget):
    # 요약 데이터 계산
    total_price = sum(money["price"] for money in money_data)
    # 저 한줄을 풀어쓴 코드
    # total_price = 0
    # for money in money_data:
    #     total_price += money["price"]
    
    # ----------------------------------------------------
    # 1. 메인 제목 영역
    # ----------------------------------------------------
    worksheet.merge_cells("A1:F1")

    title_cell = worksheet["A1"] # worksheet["A1"]: A1 셀 하나 가져오기 / worksheet[1]: 1행(row)의 모든 셀 가져오기 / worksheet["A"]: A열(column)의 모든 셀 가져오기
    title_cell.value = "Wedding Money Manager 지출 내역"
    title_cell.font = styles.font_title
    title_cell.alignment = styles.align_center

    # 병합된 A1:F1 전체에 배경색 채우기
    for col in range(1, 7): # range(시작, 끝)은 끝은 포함하지 않음. 즉, 1 ~ 6을 의미함
        worksheet.cell(row=1, column=col).fill = styles.title_fill # worksheet.cell(): 원하는 셀 객체를 가져오기 / fill: 셀의 Fill(배경색)을 지정하는 속성

    # 빈 줄 추가
    # worksheet.append([])

    # ----------------------------------------------------
    # 2. 요약 카드 영역 (A3:F4)
    # ----------------------------------------------------
    cards = [
        ("A3:B3", "A4:B4", "A3", "A4", "총 예산", budget),
        ("C3:D3", "C4:D4", "C3", "C4", "총 지출", total_price),
        ("E3:F3", "E4:F4", "E3", "E4", "남은 금액", budget - total_price),
    ] # 리스트[] 안에 튜플()이 3개 들어있는 형태

    for head_range, val_range, h_cell, v_cell, label, val in cards:
        # 셀 병합
        worksheet.merge_cells(head_range)
        worksheet.merge_cells(val_range)

        # 값 설정
        worksheet[h_cell] = label
        worksheet[v_cell] = val

        # 서식 지정
        worksheet[v_cell].number_format = '#,##0"원"'

        # 스타일 적용 (병합 셀 전체 스타일링)
        for row_idx, r_range in enumerate([head_range, val_range]): # enumerate()를 사용하면 인덱스와 값을 동시에 가져옴
            is_head = row_idx == 0

            # openpyxl에서는 worksheet["A3:B3"]처럼 범위(Range)를 지정하면 무조건 '2차원 튜플(행들의 모음)' 구조로 반환!
            for row in worksheet[r_range]: # worksheet[r_range]는 행(row) 단위의 튜플을 반환
                for cell in row: # row = (A3, B3) & cell = A3, B3
                    cell.fill = styles.card_header_fill if is_head else styles.card_val_fill
                    cell.font = styles.font_card_header if is_head else styles.font_card_val
                    cell.alignment = styles.align_center
                    cell.border = styles.border_card_header if is_head else styles.border_card_val

    worksheet.row_dimensions[3].height = 22
    worksheet.row_dimensions[4].height = 28
    
    # ----------------------------------------------------
    # 3. 테이블 헤더 영역
    # ----------------------------------------------------
    headers = ["날짜", "분류", "항목", "구매처", "금액", "결제수단"]

    worksheet.append([])  # 5행 생성을 위한 빈 레이아웃 맞춰주기 (append 활용 대신 direct 지정도 가능)

    # 직접 6행에 헤더 할당
    for col_idx, header in enumerate(headers, 1): # enumerate(..., 1)의 의미: 인덱스를 0이 아닌 1부터 시작
        cell = worksheet.cell(row=6, column=col_idx, value=header) # 6번째 행(row=6), col_idx번째 열에 header 값을 입력하고 해당 셀 객체를 가져옴
        cell.font = styles.font_table_header
        cell.fill = styles.table_header_fill
        cell.alignment = styles.align_center
        cell.border = styles.border_table_header

    worksheet.row_dimensions[6].height = 20

    # ----------------------------------------------------
    # 4. 데이터 영역
    # ----------------------------------------------------
    start_row = 7

    for money in money_data:
        worksheet.append([
            money["date"],
            money["category"],
            money["item"],
            money.get("shop", ""),
            money["price"],
            money.get("payment", "")
        ])

    end_row = worksheet.max_row

    # 데이터 행 스타일 적용
    for row in range(start_row, end_row + 1):  # range(시작, 끝)은 끝은 포함하지 않음. 즉, start_row ~ end_row을 의미함
        worksheet.row_dimensions[row].height = 22

        # 날짜, 분류, 구매처, 결제수단 -> 중앙 정렬
        for col_idx in [1, 2, 4, 6]:
            cell = worksheet.cell(row=row, column=col_idx)
            cell.alignment = styles.align_center
            cell.font = styles.font_data
            cell.border = styles.border_all

        # 항목 -> 좌측 정렬
        cell_item = worksheet.cell(row=row, column=3)
        cell_item.alignment = styles.align_left
        cell_item.font = styles.font_data
        cell_item.border = styles.border_all

        # 금액 -> 우측 정렬 및 천단위 콤마
        cell_price = worksheet.cell(row=row, column=5)
        cell_price.alignment = styles.align_right
        cell_price.font = styles.font_data
        cell_price.number_format = "#,##0"
        cell_price.border = styles.border_all

    # ----------------------------------------------------
    # 5. 기타 설정 (필터, 고정, 너비 자동 조절)
    # ----------------------------------------------------
    # 필터 적용
    # worksheet.auto_filter.ref = worksheet.dimensions # auto_filter.ref: 필터를 적용할 범위 지정 / worksheet.dimensions: 현재 데이터 영역을 자동으로 가져오는 속성 (ex. "A1:F20")
    worksheet.auto_filter.ref = f"A6:F{end_row}" # 제목이랑 요약이 추가되면서 필터 영역 수정함

    # 틀 고정(헤더 영역까지 보이고 아래 데이터 스크롤)
    worksheet.freeze_panes = "A7" # freeze_panes: 특정 위치 위쪽&왼쪽 고정. A2: A열 왼쪽은 없으니까 고정 없음 & 2행 위쪽인 1행만 고정. 즉, 2번째 행부터 스크롤 하되 1번째 행은 계속 보여달라 (ex. B2: A열 고정, 1행 고정)

    # 한글 글자 수를 고려한 칼럼 너비 자동 조절
    for col in worksheet.columns: # worksheet.columns: 열(A, B, C, D...) 가져오기
        max_len = 0
        col_letter = get_column_letter(col[0].column) # col[0]: A1, A2, A3... 중에 A1 / col[0].column: A열 -> 1, B열 -> 2 반환 / get_column_letter: 1 -> A, 2 -> B, 3 -> C...로 바꿔주는 함수
        # col[0].row: 행 번호 / col[0].column: 열 번호 / col[0].value: 데이터 값

        for cell in col:
            if cell.row < 6: # 상단 요약 카드/제목은 너비 계산에서 제외
                continue
            if cell.value is not None:
                val_str = str(cell.value)

                # 한글/전각 문자는 길이는 2, 그 외는 1로 계산
                calculated_len = sum(2 if ord(char) > 127 else 1 for char in val_str) # ord(char)는 문자의 유니코드 번호. ord(char) > 127인 경우는 알파벳, 숫자, 기본 기호를 제외한 한글, 한자, 특수문자에 해당

                # 저 한줄을 풀어쓴 코드
                # calculated_len = 0
                # for char in val_str: # 문자열에서 글자를 하나씩 꺼냄
                #     if ord(char) > 127: # 한글/특수문자 등 (유니코드 128 이상)
                #         calculated_len += 2
                #     else:  # 영문, 숫자, 기본 기호 등
                #         calculated_len += 1

                max_len = max(max_len, calculated_len)

        # 최소 14, 기본 여백 +4 추가
        # worksheet.column_dimensions[col_letter].width = max(max_len + 5, 14) # max_len이 14보다 작더라도 최소 14의 너비를 보장
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 5, 14), 30) # 열 너비가 최대 30을 넘지 않도록 잘라줌
        # worksheet.column_dimensions["A"] 이런식으로 사용