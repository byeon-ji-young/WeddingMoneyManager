# 분석(요약) 화면

import excel.styles as styles

from collections import defaultdict # defaultdict: 카테고리별 합계를 만들 때 사용
from openpyxl.chart import BarChart, Reference #BarChart: 막대 차트 객체, Reference: 어느 셀을 차트 데이터로 사용할지 지정 

def create_summary_sheet(worksheet, money_data, budget):
    # 시트 이름
    worksheet.title = "대시보드"
    
    # ----------------------------------------------------
    # 이벤트 처리 함수 정의
    # ----------------------------------------------------
    # 병합 영역 전체에 스타일(배경색, 테두리, 정렬) 적용
    def style_range(worksheet, cell_range, fill=None, font=None, alignment=None, border=styles.border_all): # fill=None: fill 값을 안 넣으면 기본적으로 None으로 처리
        for row in worksheet[cell_range]:
            for cell in row:
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                if border:
                    cell.border = border

    # 특정 셀에 스타일(배경색, 테두리, 정렬) 적용
    def style_cell(worksheet, cell_address, fill=None, font=None, alignment=None, border=styles.border_all,):
        cell = worksheet[cell_address]

        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    # 막대 그리기
    def create_bar(price, max_price, max_length=10):
        if max_price == 0:
            return ""

        length = int(price / max_price * max_length) # max_length: 막대 최대 길이 조절
        # length = max(1, length)

        return "█" * length

    # ----------------------------------------------------
    # 데이터 계산
    # ----------------------------------------------------
    # 요약 데이터
    total_price = sum(money["price"] for money in money_data)

    # 예산 사용률
    usage_rate = 0
    if budget > 0:
        usage_rate = total_price / budget * 100

    # 카테고리별 지출
    category_data = defaultdict(int) # defaultdict(int): 처음 보는 키는 자동으로 만들어줌. 기본값이 0인 딕셔너리(dictionary)를 만드는 코드. 즉, category_data["가전"] = 0을 자동으로 만들어줌
    for money in money_data:
        category_data[money["category"]] += money["price"]

    # 카테고리별 지출 높은 순 정렬
    category_data = sorted(
        category_data.items(), # items(): {category: price, category2: price2, ...} 이 형태를 [(category, price), (category2, price2), ...] 형태로 바꿔줌
        key=lambda x: x[1], # lambda: 일회성 작은 함수를 만드는 문법 / x[1]: 두 번째 값(price)을 기준으로 정렬
        reverse=True # reverse=True: 내림차순
    )

    # 최고 금액 구하기
    max_categor = 0
    if category_data:
        max_category = category_data[0][1] # 이미 내림차순으로 정렬해놓은 상태라 첫번째 리스트 뽑으면 됨

    # 지출 금액 TOP 5
    top_expenses = sorted(
        money_data,
        key=lambda x: x["price"],
        reverse=True
    )[:5] # 리스트[시작:끝]: 끝 인덱스는 포함하지 않음. 즉, [:5]는 0번째부터 5번째 전까지 가져와라

    # 결제수단별 지출
    payment_data = defaultdict(int)
    for money in money_data:
        payment = money.get("payment", "미입력") # money["payment"]: payment가 없으면 keyError 발생. get("payment", "미입력"): 있으면 가져오고 아니면 미입력
        payment_data[payment] += money["price"]

    # 결제수단별 지출 비율
    payment_total = sum(payment_data.values()) # values(): 70000, 20000, 10000, ...
    payment_ratio = {}

    for payment, price in payment_data.items(): # items(): ("카드", 70000), ("현금", 20000), ("계좌이체", 10000), ...
        if payment_total > 0:
            payment_ratio[payment] = price / payment_total * 100
        else:
            payment_ratio[payment] = 0

    # 결제수단별 최고 금액 구하기
    max_payment = max(payment_data.values()) if payment_data else 0
    
    # ----------------------------------------------------
    # 1. 메인 제목 영역
    # ----------------------------------------------------
    worksheet.merge_cells("A1:F1")

    title_cell = worksheet["A1"] # worksheet["A1"]: A1 셀 하나 가져오기 / worksheet[1]: 1행(row)의 모든 셀 가져오기 / worksheet["A"]: A열(column)의 모든 셀 가져오기
    title_cell.value = "결혼 준비 대시보드"

    style_range(worksheet, "A1:F1", fill=styles.title_fill, font=styles.font_title, alignment=styles.align_center)

    # ----------------------------------------------------
    # 2. 예산 사용률 (좌측)
    # ----------------------------------------------------
    worksheet.merge_cells("A3:C3")
    worksheet.merge_cells("A4:C7")
    worksheet.merge_cells("A8:C8")

    worksheet["A3"] = "예산 사용률"
    worksheet["A4"] = f"{usage_rate:.1f}%"
    worksheet["A8"] = f"지출 금액 {total_price:,}원, 남은 예산 {budget-total_price:,}원"

    # 제목
    style_range(worksheet, "A3:C3", fill=styles.summary_title_fill, font=styles.font_summary_title, alignment=styles.align_center)

    # 숫자
    style_range(worksheet, "A4:C7", fill=styles.summary_fill, font=styles.font_summary_value, alignment=styles.align_center)

    # 설명
    style_range(worksheet, "A8:C8", fill=styles.summary_fill, font=styles.font_summary_sub, alignment=styles.align_center)

    # ----------------------------------------------------
    # 3. 카테고리별 지출 (우측)
    # ----------------------------------------------------
    worksheet.merge_cells("D3:F3")
    worksheet["D3"] = "카테고리별 지출"

    style_range(worksheet, "D3:F3", fill=styles.summary_title_fill, font=styles.font_summary_title, alignment=styles.align_center)

    row = 4
    for category, price in category_data[:5]:
        bar = create_bar(price, max_category)
        
        worksheet[f"D{row}"] = category
        worksheet[f"E{row}"] = bar
        worksheet[f"F{row}"] = price
        worksheet[f"F{row}"].number_format = '#,##0"원"'

        style_range(worksheet, f"D{row}:E{row}", font=styles.font_data, alignment=styles.align_left)
        style_cell(worksheet, f"F{row}", font=styles.font_data, alignment=styles.align_right)

        worksheet.row_dimensions[row].height = 20

        row += 1

    while row <= 8: # 4부터 5개니까 4,5,6,7,8이라서 8로 조건을 줌
        worksheet.merge_cells(f"D{row}:E{row}")
        style_range(worksheet, f"D{row}:F{row}")

        row += 1

    # ----------------------------------------------------
    # 4. 최대 지출 TOP 5 (좌측)
    # ----------------------------------------------------
    worksheet.merge_cells("A9:C9")
    worksheet["A9"] = "최대 지출 TOP 5"
    style_range(worksheet, "A9:C9", fill=styles.summary_title_fill, font=styles.font_summary_title, alignment=styles.align_center)

    # 서브 헤더
    worksheet.merge_cells("A10:B10")
    worksheet["A10"] = "항목"
    worksheet["C10"] = "금액"
    style_range(worksheet, "A10:C10", fill=styles.summary_fill, font=styles.font_card_header, alignment=styles.align_center)

    row_top = 11
    for money in top_expenses:
        worksheet.merge_cells(f"A{row_top}:B{row_top}")
        worksheet[f"A{row_top}"] = money["item"]
        worksheet[f"C{row_top}"] = money["price"]
        worksheet[f"C{row_top}"].number_format = '#,##0"원"'

        style_range(worksheet, f"A{row_top}:B{row_top}", font=styles.font_data, alignment=styles.align_left)
        style_cell(worksheet, f"C{row_top}", font=styles.font_data, alignment=styles.align_right)

        worksheet.row_dimensions[row_top].height = 20

        row_top += 1
    
    # ----------------------------------------------------
    # 5. 결제수단 분석 (우측)
    # ----------------------------------------------------
    worksheet.merge_cells("D9:F9")
    worksheet["D9"] = "결제수단별 금액"
    style_range(worksheet, "D9:F9", fill=styles.summary_title_fill, font=styles.font_summary_title, alignment=styles.align_center)

    # 서브 헤더
    worksheet.merge_cells("D10:E10")
    worksheet["D10"] = "결제수단"
    worksheet["F10"] = "금액"
    style_range(worksheet, "D10:F10", fill=styles.summary_fill, font=styles.font_card_header, alignment=styles.align_center)

    row_pay = 11
    for payment, price in payment_data.items():
        bar = create_bar(price, payment_total)

        ratio = price / payment_total * 100 if payment_total else 0

        worksheet[f"D{row_pay}"] = payment
        worksheet[f"E{row_pay}"] = bar
        worksheet[f"F{row_pay}"] = f"{price:,}원({ratio:.1f}%)"
        # worksheet[f"F{row_pay}"].number_format = '#,##0"원"'

        style_range(worksheet, f"D{row_pay}:D{row_pay}", font=styles.font_data, alignment=styles.align_center)
        style_cell(worksheet, f"E{row_pay}", font=styles.font_data, alignment=styles.align_left)
        style_cell(worksheet, f"F{row_pay}", font=styles.font_data, alignment=styles.align_right)

        worksheet.row_dimensions[row_pay].height = 20

        row_pay += 1

    # ----------------------------------------------------
    # 6. 차트 데이터
    # ----------------------------------------------------
    chart_row = 2
    
    for category, price in category_data:
        worksheet[f"H{chart_row}"] = category
        worksheet[f"I{chart_row}"] = price

        chart_row += 1

    CATEGORY_COL = 8 # H열
    PRICE_COL = 9 # I열

    chart = BarChart() # 막대그래프 객체 생성

    # 데이터 범위 지정 - 세로축(Y축) 데이터
    data = Reference(worksheet, min_col=PRICE_COL, min_row=2, max_row=chart_row - 1) # Reference: 차트가 사용할 셀 범위 / min_col=9: 9번째 열(I열)

    # 카테고리 범위 지정 - 가로축(X축)
    categories = Reference(worksheet, min_col=CATEGORY_COL, min_row=2, max_row=chart_row - 1)

    # 차트에 연결 (데이터 추가, 카테고리 추가)
    chart.add_data(data)
    chart.set_categories(categories)

    # 차트 꾸미기
    chart.title = "카테고리별 지출"
    chart.width = 13
    chart.height = 7

    # 시트에 추가
    worksheet.add_chart(chart, "A18") # A18: 왼쪽 위 시작 위치. 즉, A18부터 그래프가 그려짐

    # ----------------------------------------------------
    # 열 너비
    # ----------------------------------------------------
    # worksheet.column_dimensions["A"].width = 18
    for col in ["A", "B", "C", "D", "E", "F"]:
        worksheet.column_dimensions[col].width = 18