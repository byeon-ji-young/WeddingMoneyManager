# 공통 디자인

from openpyxl.styles import Font, PatternFill, Alignment # Font: 글자 스타일, PatternFill: 셀 배경색, Alignment: 정렬
from openpyxl.styles import Border, Side

# ----------------------------------------------------
# Color
# ----------------------------------------------------
color_primary = "1E293B"  # 메인 헤더 / 제목 (진한 슬레이트)
color_accent = "3B82F6"  # 요약 카드 포인트 (블루)
color_header_bg = "F1F5F9"  # 테이블 헤더 (연한 회색)
color_card_bg = "F8FAFC"  # 요약 카드 배경

# ----------------------------------------------------
# Fonts
# ----------------------------------------------------
font_title = Font(name="맑은 고딕", size=15, bold=True, color="FFFFFF")
font_card_header = Font(name="맑은 고딕", size=10, bold=True, color="475569")
font_card_val = Font(name="맑은 고딕", size=13, bold=True, color="0F172A")
font_card_val_danger = Font(name="맑은 고딕", size=13, bold=True, color="FF0000")
font_table_header = Font(name="맑은 고딕", size=10, bold=True, color="1E293B")
font_data = Font(name="맑은 고딕", size=10, color="334155")
font_summary_title = Font(name="맑은 고딕", size=10, bold=True, color="475569")
font_summary_value = Font(name="맑은 고딕", size=18, bold=True, color="0F172A")
font_summary_sub = Font(name="맑은 고딕", size=10, color="64748B")
font_summary_value_danger = Font(name="맑은 고딕", size=24, bold=True, color="FF0000")
font_summary_sub_danger = Font(name="맑은 고딕", size=10, bold=True, color="FF0000")

# ----------------------------------------------------
# Alignments
# ----------------------------------------------------
align_center = Alignment(horizontal="center", vertical="center") # horizontalL 가로 정렬, vertical: 세로 정렬
align_right = Alignment(horizontal="right", vertical="center", indent=1) # Indent: 좌측/우측 텍스트 여백 ex. indent=1
align_left = Alignment(horizontal="left", vertical="center", indent=1)

# ----------------------------------------------------
# Borders
# ----------------------------------------------------
thin_side = Side(style="thin", color="CBD5E1") # Side: 테두리 한쪽의 스타일
thick_bottom_side = Side(style="medium", color="94A3B8")

border_all = Border(
    left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
)
border_card_header = Border(
    left=thin_side, right=thin_side, top=thin_side, bottom=Side(style="none")
)
border_card_val = Border(
    left=thin_side, right=thin_side, top=Side(style="none"), bottom=thin_side
)
border_table_header = Border(
    left=thin_side,
    right=thin_side,
    # top=Side(style="none"),
    top=thin_side,
    bottom=thin_side
)
summary_border = Border(
    left=thin_side,
    right=thin_side,
    top=thin_side,
    bottom=thin_side
)

# ----------------------------------------------------
# Fill
# ----------------------------------------------------
title_fill = PatternFill( # PatternFill: 배경 채우기(Fill) 스타일을 만드는 클래스
    fill_type="solid", # solid: 단색
    start_color=color_primary, 
    end_color=color_primary
)
card_header_fill = PatternFill(
    fill_type="solid",
    start_color="E2E8F0"
)
card_val_fill = PatternFill(
    fill_type="solid",
    start_color=color_card_bg
)
table_header_fill = PatternFill(
    fill_type="solid",
    start_color=color_header_bg
)
table_zebra_fill = PatternFill(
    start_color="F9FAFB", 
    end_color="F9FAFB", 
    fill_type="solid"
)
summary_fill = PatternFill(
    fill_type="solid",
    start_color="F8FAFC"
)
summary_title_fill = PatternFill(
    fill_type="solid",
    start_color="E2E8F0"
)