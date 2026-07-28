import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment # Font: 글자 스타일, PatternFill: 셀 배경색, Alignment: 정렬
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

from datetime import datetime

today = datetime.now().strftime("%Y-%m-%d")

# 상태(State)를 가지고 있으면 Class & 기능만 수행하면 Function. 즉, 기능이 많아지고 상태를 관리해야 하면 Class.
def export_excel(money_data):
    # 엑셀파일
        workbook = Workbook() # workbook(): 엑셀 파일 자체를 만드는 객체
    
        # 엑셀 안의 첫 번째 시트
        worksheet = workbook.active # active: 현재 선택되어 있는 시트 가져오기
    
        # 시트 이름 변경
        worksheet.title = "WeddingMoney"
    
        # 헤더
        worksheet.append([
            "날짜",
            "분류",
            "항목",
            "구매처",
            "금액",
            "결제수단"
        ])

        # 헤더 스타일 적용
        for cell in worksheet[1]: # worksheet[1]: 1행 전체를 가져오기
            # 글자
            cell.font = Font(
                bold=True,
                color="FFFFFF",
                # size=12
            )

            # 배경색
            cell.fill = PatternFill(
                fill_type="solid", # solid: 단색
                start_color="4F81BD"
            )

            # 가운데 정렬
            cell.alignment = Alignment(
                horizontal="center", # 가로 중앙 정렬
                vertical="center" # 세로 중앙 정렬
            )

        # 데이터
        for money in money_data:
            worksheet.append([
                money["date"],
                money["category"],
                money["item"],
                money.get("shop", ""),
                money["price"],
                money.get("payment", "")
            ])

        # 엑셀 필터 적용
        worksheet.auto_filter.ref = worksheet.dimensions # auto_filter.ref: 필터를 적용할 범위 지정 / worksheet.dimensions: 현재 데이터 영역을 자동으로 가져오는 속성 (ex. "A1:F20")

        # 첫 번째 행 고정
        worksheet.freeze_panes = "A2" # freeze_panes: 특정 위치 위쪽&왼쪽 고정
        # A2를 한 이유: A열 왼쪽은 없으니까 고정 없음 & 2행 위쪽인 1행만 고정. 즉, 2번째 행부터 스크롤 하되 1번째 행은 계속 보여달라 (ex. B2: A열 고정, 1행 고정)

        # 전체 셀 테두리 적용
        thin_border = Border(
            left=Side(style="thin"), # Side: 테두리 한쪽의 스타일
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                
        # 금액 컬럼 숫자 표시 형식 적용
        # for cell in worksheet["E"]: # E열 전체 가져오기. 이렇게 하면 E1 = 금액까지 전부 포함.
        #     cell.number_format = '#,##0'
        for row in range(2, worksheet.max_row + 1):
            worksheet[f"E{row}"].number_format = '#,##0'

        # 열 너비 자동 조절
        for column in worksheet.columns: # worksheet.columns: 열(A, B, C, D...) 가져오기
            max_length = 0
            
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            column_letter = get_column_letter(column[0].column) # column[0]: A1, A2, A3... 중에 A1 / column[0].column: A열 -> 1, B열 -> 2 반환 / get_column_letter: 1 -> A, 2 -> B, 3 -> C...로 바꿔주는 함수
            # column[0].row: 행 번호 / column[0].column: 열 번호 / column[0].value: 데이터 값

            worksheet.column_dimensions[column_letter].width =min(max(max_length * 1.5, 12), 30)
            # worksheet.column_dimensions["A"] 이런식으로 사용

            print(column_letter, max_length + 3)

        # Excel 저장
        filename = f"WeddingMoney_{today}.xlsx"
        workbook.save(filename)
    
        # Excel 자동 실행
        os.startfile(filename)